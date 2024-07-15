import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook,load_workbook

def formatar_data(event):
    # Remove todos os caracteres que não sejam números
    input_value = event.widget.get().replace('/', '').replace('-', '')

    # Limita o tamanho do texto inserido para 8 caracteres (ddmmaaaa)
    input_value = input_value[:8]

    # Adiciona as barras conforme o formato dd/mm/aaaa
    if len(input_value) > 2:
        input_value = input_value[:2] + '/' + input_value[2:]
    if len(input_value) > 5:
        input_value = input_value[:5] + '/' + input_value[5:]

    # Atualiza o valor do campo de entrada
    event.widget.delete(0, tk.END)
    event.widget.insert(tk.END, input_value)

def callback(event):
    x = event.widget.winfo_pointerx() - event.widget.winfo_rootx()
    y = event.widget.winfo_pointery() - event.widget.winfo_rooty()
    print("Coordenada X dentro do widget:", x)
    print("Coordenada Y dentro do widget:", y)

def show_placeholder(event):
    current_widget = event.widget
    current_value = current_widget.get()
    if current_value == "":
        current_widget.delete(0, tk.END)
        current_widget.insert(0, current_widget.placeholder)
        current_widget.config(fg="gray")

def restore_placeholder(event):
    current_widget = event.widget
    current_value = current_widget.get()
    if current_value == current_widget.placeholder:
        current_widget.delete(0, tk.END)
        current_widget.config(fg="black")


def on_entry_focus_in(event):
    entry = event.widget
    if entry.get() == f"Participante {entry.index('end')}" or entry.get() == f"Por que {entry.index('end')}":
        entry.delete(0, tk.END)
        entry.config(fg="black")

def on_entry_focus_out(event):
    entry = event.widget
    if entry.get() == "":
        entry.insert(0, f"Participante {entry.index('end')}")
        entry.config(fg="gray")

def focus_next_widget(event):
    event.widget.tk_focusNext().focus()
    return("break")

def encontrar_proxima_linha_vazia(sheet, coluna_referencia=1):
    # Começando na linha 2, verificando cada linha até encontrar uma vazia
    for row in range(2, sheet.max_row + 1):
        if not sheet.cell(row=row, column=coluna_referencia).value:
            return row
    # Se não houver linha vazia, retorna a próxima linha após a última linha preenchida
    return sheet.max_row + 1

def encontrar_ultimo_numero():
    try:
        wb = load_workbook('RQ-CNC_Ed01.xlsx')
        sheet = wb.active
        ultimo_numero = 0  # Inicializa com 0
        # Procura na coluna de números e pega o último valor não vazio e numérico
        for cell in sheet['A'][::-1]:
            if cell.value and cell.value.replace('/', '').isdigit():
                ultimo_numero = int(cell.value.split('/')[0])
                break  # Interrompe a busca quando o último número válido for encontrado
        return ultimo_numero
    except FileNotFoundError:
        # Retorna 0 se o arquivo não existir ou estiver vazio
        return 0

def apagar_tudo():
    # Limpar todas as entradas
    date_entry.delete(0, tk.END)
    responsavel_entry.delete(0, tk.END)
    tipo_menu.set("")
    acao_menu.set("")
    conclusao_menu.set("")
    aberto_entry.delete(0, tk.END)
    encerrado_entry.delete(0, tk.END)
    setor_entry.delete(0, tk.END)
    tipoot_entry.delete(0, tk.END)
    documento_ref_entry.delete(0, tk.END)
    assunto_text.delete("1.0", tk.END)
    descricao_text.delete("1.0", tk.END)
    mudancas_input.delete(0, tk.END)
    falha_input.delete(0, tk.END)
    coleta_combobox.set("")
    comunicadas_combobox.set("")
    tipoacao_combobox.set("")
    observacoes_input.delete("1.0", tk.END)
    brainstorming_input.delete(0, tk.END)
    causa_efeito_input.delete(0, tk.END)
    cinco_porques_input.delete(0, tk.END)
    causas_basicas_input.delete(0, tk.END)

    for i in range(1, 6):
        entry_widgets[f"part{i}"].delete(0, tk.END)
        entry_widgets[f"part{i}"].insert(tk.END, entry_widgets[f"part{i}"].placeholder)
        entry_widgets[f"part{i}"].config(fg="gray")

    for i in range(1, 6):
        entrypq_widgets[f"part0{i}"].delete(0, tk.END)
        entrypq_widgets[f"part0{i}"].insert(tk.END, entrypq_widgets[f"part0{i}"].placeholder)
        entrypq_widgets[f"part0{i}"].config(fg="gray")

    for row in range(1,6):
        for col in range(5):
            entradas[(row, col)].delete(0, tk.END)

# Define o valor do combobox como vazio uma vez antes de começar a iterar sobre as linhas
    for row in range(5, 9):

        for col in range(9):
            if (row, col) in entradas1:  # Verifica se o widget existe no dicionário
                widget = entradas1[(row, col)]
                if isinstance(widget, ttk.Combobox):  # Verifica se é um combobox
                    widget.set("")  # Define o valor do combobox como vazio
                else:  # Para outros tipos de widgets
                    widget.delete(0, tk.END) if isinstance(widget, tk.Entry) else widget.delete("1.0", tk.END)



        for widget in entradas1.values():
            if isinstance(widget, ttk.Combobox):
                widget.set("")


                



    comunicadas_input.set("")  # Define o valor da StringVar como uma string vazia para limpar o conteúdo
    testes_input.delete(0, tk.END)
    coleta_combobox_2.set("")
    problema_resolvido_input.delete("1.0", tk.END)
    conclusao_menu4.set("")

    for i in range(1, 5):
        entry_widgets4[f"part{i}"].delete(0, tk.END)
        entry_widgets4[f"part{i}"].insert(tk.END, entry_widgets4[f"part{i}"].placeholder)
        entry_widgets4[f"part{i}"].config(fg="gray")

    analise_encerramento_input.delete("1.0", tk.END)


def salvar_excel():
    # Carregando ou criando um arquivo Excel
    try:
        wb = load_workbook('RQ-CNC_Ed01.xlsx')
    except FileNotFoundError:
        wb = Workbook()
    sheet = wb.active

    ultimo_numero = encontrar_ultimo_numero() + 1
    novo_numero = f"{ultimo_numero:03d}/{2024}"  # Formata o número com três dígitos e o ano



    # Definindo os cabeçalhos das colunas, se necessário
    if True:
        cabecalhos = ['Número', 'Data', 'Responsável', 'Não-Conformidade','Oportunidade de Melhoria', 'Corretiva','Preventiva','Melhoria', 'Eficaz','Ineficaz', 
                      'Aberto em', 'Encerrado em', 'Setor', 'Tipo', 'Documento Ref.', 
                      'Assunto', 'Participantes', 'Descrição do Problema', '5 Por ques',
                      
                      'Item','Atividade','Responsável','Prazo Inicial', 'Prazo Final', 
                      'Não','Sim','Sim Infos', 
                      'Não','Sim','Sim Infos',
                      'Não','Sim','Folha de Verificação','Histograma','Diagrama de Pareto','Diagrama de Causa Raiz','Outro','Info Outro',
                      'Observções',
                      'Não','Sim','Anexo',
                      'Não','Sim','Anexo', 
                      'Não','Sim','Anexo', 
                      'Causas Básicas',

                      'Corretiva','Preventiva','Melhoria',
                      'Não','Sim','Como',
                      'Item','Tipo','Descrição da Ação','Responsável','Prazo Incial','Prazo Final',
                      'Não','Sim','Como',
                      'Não','Sim','Folha de Verificação','Histograma','Diagrama de Pareto','Diagrama de Causa Raiz','Outro','Info Outro',
                      
                      'O Problema foi Resolvido?', 
                      'Eficaz','Ineficaz','Ação', 
                      'Participantes', 'Análise de Encerramento']
        for col, cabecalho in enumerate(cabecalhos, start=1):
            sheet.cell(row=1, column=col).value = cabecalho


    # Encontrando a próxima linha vazia
    proxima_linha = encontrar_proxima_linha_vazia(sheet)

    coleta_nao = ''
    coleta_sim = ''
    coleta_fv = ''
    coleta_his = ''
    coleta_dp = ''
    coleta_dcr = ''
    coleta_outro = ''
    coleta_info = ''

    tipo_value = ''
    tipo1_value = ''
    acao_corretiva = ''
    acao_prev = ''
    acao_mel = ''
    conclusao_efi = ''
    conclusao_ine = ''
    mudancas_sim = ''
    mudancas_nao = ''
    mudancas_info = ''
    falha_sim = ''
    falha_nao = ''
    falha_info = ''

    brain_sim = ''
    brain_nao = ''
    brain_info = ''
    diagramace_sim = ''
    diagramace_nao = ''
    diagramace_info = ''
    tecporq_sim = ''
    tecporq_nao = ''
    tecporq_info = ''

    tipoacao_cor = ''
    tipoacao_prev = ''
    tipoacao_mel = ''
    comunicadas_sim = ''
    comunicadas_nao = ''
    comunicadas_como = ''

    teste_sim = ''
    teste_nao = ''
    teste_como = ''

    pag3_coleta_nao = ''
    pag3_coleta_sim = ''
    pag3_coleta_fv = ''
    pag3_coleta_his = ''
    pag3_coleta_dp = ''
    pag3_coleta_dcr = ''
    pag3_coleta_outro = ''
    pag3_coleta_info = ''

    pag4_conclusao_efi = ''
    pag4_conclusao_ine = ''
    pag4_conclusao_acao = ''


    # Defina os valores com base nas seleções do menu
    pag4_conclusao_value = conclusao_menu4.get()
    if pag4_conclusao_value == 'Eficaz':
        pag4_conclusao_efi = 'X'
        pag4_conclusao_ine = ''
        pag4_conclusao_acao = ''
    elif pag4_conclusao_value == 'Ineficaz':
        pag4_conclusao_efi = ''
        pag4_conclusao_ine = 'X'
        pag4_conclusao_acao = ''
    elif pag4_conclusao_value:
        pag4_conclusao_efi = ''
        pag4_conclusao_ine = ''
        pag4_conclusao_acao = pag4_conclusao_value

    
    coleta2_value = coleta_combobox_2.get()
    if coleta2_value == 'Não':
        pag3_coleta_nao = 'X'
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Sim':
        pag3_coleta_nao = ''
        pag3_coleta_sim = 'X'
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Folha de Verificação':
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = 'X'
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Histograma':
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = 'X'
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Diagrama de Pareto':
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = 'X'
        pag3_coleta_dcr = ''
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Diagrama de Causa Raiz':
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = 'X'
        pag3_coleta_outro = ''
        pag3_coleta_info = ''
    elif coleta2_value == 'Outros':
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = 'X'
        pag3_coleta_info = ''
    elif coleta2_value:
        pag3_coleta_nao = ''
        pag3_coleta_sim = ''
        pag3_coleta_fv = ''
        pag3_coleta_his = ''
        pag3_coleta_dp = ''
        pag3_coleta_dcr = ''
        pag3_coleta_outro = 'X'
        pag3_coleta_info = coleta2_value



    teste_value = testes_input.get()
    if teste_value == 'Sim':
        teste_sim = 'X'
        teste_nao = ''
        teste_como = ''
    elif teste_value == 'Não':
        teste_sim = ''
        teste_nao = 'X'
        teste_como = ''
    elif teste_value:
        teste_sim = 'X'
        teste_nao = ''
        teste_como =teste_value


    comunicadas_value = comunicadas_input.get()
    if comunicadas_value == 'Sim':
        comunicadas_sim = 'X'
        comunicadas_nao = ''
        comunicadas_como = ''
    elif comunicadas_value == 'Não':
        comunicadas_sim = ''
        comunicadas_nao = 'X'
        comunicadas_como = ''
    elif comunicadas_value:
        comunicadas_sim = 'X'
        comunicadas_nao = ''
        comunicadas_como = comunicadas_value




    tipo_acao_value = tipoacao_combobox.get()
    if tipo_acao_value == 'Corretiva':
        tipoacao_cor = 'X'
        tipoacao_prev = ''
        tipoacao_mel = ''
    elif tipo_acao_value == 'Preventiva':
        tipoacao_cor = ''
        tipoacao_prev = 'X'
        tipoacao_mel = ''
    elif tipo_acao_value == 'Melhoria':
        tipoacao_cor = ''
        tipoacao_prev = ''
        tipoacao_mel = 'X'

    tecporq_value = cinco_porques_input.get()
    if tecporq_value == 'Não':
        tecporq_sim = ''
        tecporq_nao = 'X'
        tecporq_info = ''
    elif tecporq_value == 'Sim':
        tecporq_sim = 'X'
        tecporq_nao = ''
        tecporq_info = ''
    elif tecporq_value:
        tecporq_sim = 'X'
        tecporq_nao = ''
        tecporq_info = tecporq_value



    diagrama_value = causa_efeito_input.get()
    if diagrama_value == 'Não':
        diagramace_sim = ''
        diagramace_nao = 'X'
        diagramace_info = ''
    elif diagrama_value == 'Sim':
        diagramace_sim = 'X'
        diagramace_nao = ''
        diagramace_info = ''
    elif diagrama_value:
        diagramace_sim = 'X'
        diagramace_nao = ''
        diagramace_info = diagrama_value


    brain_value = brainstorming_input.get()
    if brain_value == 'Não':
        brain_sim = ''
        brain_nao = 'X'
        brain_info = ''
    elif brain_value == 'Sim':
        brain_sim = 'X'
        brain_nao = ''
        brain_info = ''
    elif brain_value:
        brain_sim = 'X'
        brain_nao = ''
        brain_info = brain_value



    coleta_value = coleta_combobox.get()
    if coleta_value == 'Não':
        coleta_nao = 'X'
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Sim':
        coleta_nao = ''
        coleta_sim = 'X'
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Folha de Verificação':
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = 'X'
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Histograma':
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = 'X'
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Diagrama de Pareto':
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = 'X'
        coleta_dcr = ''
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Diagrama de Causa Raiz':
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = 'X'
        coleta_outro = ''
        coleta_info = ''
    elif coleta_value == 'Outro':
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = 'X'
        coleta_info = ''
    elif coleta_value:
        coleta_nao = ''
        coleta_sim = ''
        coleta_fv = ''
        coleta_his = ''
        coleta_dp = ''
        coleta_dcr = ''
        coleta_outro = 'X'
        coleta_info = coleta_value

    tipo_menu_value = tipo_menu.get()
    if tipo_menu_value == 'Não-Conformidade':
        tipo_value = 'X'
    elif tipo_menu_value == 'Oportunidade de Melhoria':
        tipo1_value = 'X'

    acao_menu_value = acao_menu.get()
    if acao_menu_value == 'Corretiva':
        acao_corretiva = 'X'
    elif acao_menu_value == 'Preventiva':
        acao_prev = 'X'
    elif acao_menu_value == 'Melhoria':
        acao_mel = 'X'

    conclusao_menu_value = conclusao_menu.get()
    if conclusao_menu_value == 'Eficaz':
        conclusao_efi = 'X'
    elif conclusao_menu_value == 'Ineficaz':
        conclusao_ine = 'X'

    mudancas_value = mudancas_input.get()
    if mudancas_value == 'Não':
        mudancas_sim = ''
        mudancas_nao = 'X'
        mudancas_info = ''
    elif mudancas_value == 'Sim':
        mudancas_sim = 'X'
        mudancas_nao = ''
        mudancas_info = ''
    elif mudancas_value:
        mudancas_sim = 'X'
        mudancas_nao = ''
        mudancas_info = mudancas_value 

    falha_value = falha_input.get()
    if falha_value == 'Não':
        falha_sim = ''
        falha_nao = 'X'
        falha_info = ''
    elif falha_value == 'Sim':
        falha_sim = 'X'
        falha_nao = ''
        falha_info = ''
    elif falha_value:
        falha_sim = 'X'
        falha_nao = ''
        falha_info = falha_value

    # Coletando os dados das entradas
    dados = [
        #Pág1
        novo_numero,
        date_entry.get(),
        responsavel_entry.get(),
        tipo_value,
        tipo1_value,
        acao_corretiva,
        acao_prev,
        acao_mel,
        conclusao_efi,
        conclusao_ine,
        aberto_entry.get(),
        encerrado_entry.get(),
        setor_entry.get(),
        tipoot_entry.get(),
        documento_ref_entry.get(),
        assunto_text.get("1.0", tk.END).strip(),  # Obtendo o texto do widget Text
         '\n '.join([f"{entry.get()}," if entry.get() not in ["Participante 1", "Participante 2", "Participante 3", "Participante 4", "Participante 5"] else "" for entry in entry_widgets.values()]),  # Juntando os participantes com parênteses
        descricao_text.get("1.0", tk.END).strip(),  # Obtendo o texto do widget Text
        '\n '.join([f"{entry.get()}," if entry.get() not in ["Porque 1", "Porque 2", "Porque 3", "Porque 4", "Porque 5"] else "" for entry in entrypq_widgets.values()]),
        
        #Pág2
        '\n'.join([f'{entradas[(row, 0)].get()}' for row in range(1, 6)]),
        '\n'.join([f'{entradas[(row, 1)].get()}' for row in range(1, 6)]),
        '\n'.join([f'{entradas[(row, 2)].get()}' for row in range(1, 6)]),
        '\n'.join([f'{entradas[(row, 3)].get()}' for row in range(1, 6)]),
        '\n'.join([f'{entradas[(row, 4)].get()}' for row in range(1, 6)]),
        mudancas_nao,
        mudancas_sim,
        mudancas_info,
        falha_nao,
        falha_sim,
        falha_info,
        coleta_nao,
        coleta_sim,
        coleta_fv,
        coleta_his,
        coleta_dp,
        coleta_dcr,
        coleta_outro,
        coleta_info,
        observacoes_input.get("1.0", tk.END).strip(),
        brain_nao,
        brain_sim,
        brain_info,
        diagramace_nao,
        diagramace_sim,
        diagramace_info,
        tecporq_nao,
        tecporq_sim,
        tecporq_info,
        causas_basicas_input.get(),

        #Pág3
        tipoacao_cor,
        tipoacao_prev,
        tipoacao_mel,
        comunicadas_nao,
        comunicadas_sim,
        comunicadas_como,
        '\n'.join([f'{entradas1[(row, 0)].get()}' for row in range(5, 9)]),
        '\n'.join([f'{entradas1[(row, 1)].get()}' for row in range(5, 9)]),
        '\n'.join([f'{entradas1[(row, 2)].get("1.0", tk.END)}' for row in range(5, 9)]),
        '\n'.join([f'{entradas1[(row, 3)].get()}' for row in range(5, 9)]),
        '\n'.join([f'{entradas1[(row, 4)].get()}' for row in range(5, 9)]),
        '\n'.join([f'{entradas1[(row, 5)].get()}' for row in range(5, 9)]),
        teste_nao,
        teste_sim,
        teste_como,
        pag3_coleta_nao,
        pag3_coleta_sim,
        pag3_coleta_fv,
        pag3_coleta_his,
        pag3_coleta_dp,
        pag3_coleta_dcr,
        pag3_coleta_outro,
        pag3_coleta_info,

        #Pág4
        problema_resolvido_input.get("1.0", tk.END).strip(),
        pag4_conclusao_efi,
        pag4_conclusao_ine,
        pag4_conclusao_acao,
        '\n '.join([f"{entry.get()}," if entry.get() not in ["Participante 1", "Participante 2", "Participante 3", "Participante 4", "Participante 5"] else "" for entry in entry_widgets4.values()]), 
        analise_encerramento_input.get("1.0", tk.END).strip()
    ]

    # Escrevendo os dados nas células correspondentes na próxima linha vazia
    for col, dado in enumerate(dados, start=1):
        sheet.cell(row=proxima_linha, column=col).value = dado

    # Salvando o arquivo Excel
    wb.save('RQ-CNC_Ed01.xlsx')
    messagebox.showinfo("Sucesso", "Os dados foram salvos com sucesso no arquivo 'RQ-CNC_Ed01.xlsx'.")



root = tk.Tk()
root.title("Registro")

# Criando as abas
tab_control = ttk.Notebook(root)
tab1 = tk.Frame(tab_control)
tab_control.add(tab1, text='Página 1', )
tab_control.pack(expand=1, fill='both')  # Adiciona o tab_control à janela principal


# Criando a Página 1
main_frame = tk.Frame(tab1,  width=100, height=150)
main_frame.pack()

# Pagina 1 - Parte 1
part1_frame = tk.Frame(main_frame, padx=10, pady=10)
part1_frame.grid(row=0, column=0, sticky="nsew")

# tk.Label(part1_frame, text="Número:").grid(row=0, column=0, sticky="w")
# numero_entry = tk.Entry(part1_frame)
# numero_entry.grid(row=0, column=1, padx=5, pady=2, sticky="w")



tk.Label(part1_frame, text="Data:").grid(row=1, column=0, sticky="w")
date_entry = tk.Entry(part1_frame)
date_entry.grid(row=1, column=1, padx=5, sticky="w")
date_entry.bind("<KeyRelease>", formatar_data)
# date_entry.place(x=60, y=40, width=80, height=25)

tk.Label(part1_frame, text="Responsável:").grid(row=2, column=0, sticky="w")
responsavel_entry = tk.Entry(part1_frame)
responsavel_entry.grid(row=2, column=1, padx=5, sticky="w")

tk.Label(part1_frame, text="Tipo:").grid(row=3, column=0, sticky="w")
tipo_var = tk.StringVar()
tipo_var.set("")
tipo_menu = ttk.Combobox(part1_frame, values=["", "Não-Conformidade", "Oportunidade de Melhoria"], textvariable=tipo_var)
tipo_menu.grid(row=3, column=1, padx=5, sticky="w")


tk.Label(part1_frame, text="Ação:").grid(row=4, column=0, sticky="w")
acao_var = tk.StringVar()
acao_var.set("")
acao_menu =ttk.Combobox(part1_frame, values=["", "Corretiva", "Preventiva", "Melhoria"])
acao_menu.grid(row=4, column=1, padx=5, sticky="w")

tk.Label(part1_frame, text="Conclusão:").grid(row=5, column=0, sticky="w")
conclusao_var = tk.StringVar()
conclusao_var.set("")
conclusao_menu = ttk.Combobox(part1_frame, values=["", "Eficaz", "Ineficaz"])
conclusao_menu.grid(row=5, column=1, padx=5, sticky="w")

tk.Label(part1_frame, text="Aberto em:").grid(row=6, column=0, sticky="w")
aberto_entry = tk.Entry(part1_frame)
aberto_entry.grid(row=6, column=1, padx=5, sticky="w")
aberto_entry.bind("<KeyRelease>", formatar_data)

tk.Label(part1_frame, text="Encerrado em:").grid(row=7, column=0, sticky="w")
encerrado_entry = tk.Entry(part1_frame)
encerrado_entry.grid(row=7, column=1, padx=5, sticky="w")
encerrado_entry.bind("<KeyRelease>", formatar_data)

tk.Label(part1_frame, text="Setor:").grid(row=8, column=0, sticky="w")
setor_entry = tk.Entry(part1_frame)
setor_entry.grid(row=8, column=1, padx=5, sticky="w")


tk.Label(part1_frame, text="Tipo:").grid(row=9, column=0, sticky="w")
tipoot_entry = tk.Entry(part1_frame)
tipoot_entry.grid(row=9, column=1, padx=5, sticky="w") 

tk.Label(part1_frame, text="Documento Ref.:").grid(row=10, column=0, sticky="w")
documento_ref_entry = tk.Entry(part1_frame)
documento_ref_entry.grid(row=10, column=1, padx=5, sticky="w")

tk.Label(part1_frame, text="Assunto:").grid(row=11, column=0, sticky="w")
assunto_text = tk.Text(part1_frame,  width=50, height=15, font=("Arial", 9))
assunto_text.grid(row=11, column=1, padx=5, pady=5, sticky="w")
assunto_text.bind("<Tab>", focus_next_widget)

# Pagina 1 - Parte 2
part2_frame = tk.Frame(main_frame, width=400, height=800, padx=10, pady=10)
part2_frame.grid(row=0, column=1, sticky="nsew")

tk.Label(part2_frame, text="Participantes:").grid(row=1, column=0, sticky="w")
entry_widgets = {}
for i in range(1, 6):
    entry_widgets[f"part{i}"] = tk.Entry(part2_frame,width=30, fg="gray", justify="left", highlightthickness=0)
    entry_widgets[f"part{i}"].placeholder = f"Participante {i}"
    entry_widgets[f"part{i}"].insert(tk.END, entry_widgets[f"part{i}"].placeholder)
    entry_widgets[f"part{i}"].bind("<FocusIn>", restore_placeholder)
    entry_widgets[f"part{i}"].bind("<FocusOut>", show_placeholder)
    entry_widgets[f"part{i}"].grid(row=i, column=1, padx=5, pady=2, sticky="w")

tk.Label(part2_frame, text="Descrição do Problema:").grid(row=8, column=0, sticky="w")
descricao_text = tk.Text(part2_frame, width=50, height=15, font=("Arial", 9))
descricao_text.grid(row=8, column=1, padx=5, pady=5, sticky="w")
descricao_text.bind("<Tab>", focus_next_widget)

tk.Label(part2_frame, text="Por ques:").grid(row=9, column=0, sticky="w")
entrypq_widgets = {}
for i in range(1, 6):
    entrypq_widgets[f"part0{i}"] = tk.Entry(part2_frame,width=60, fg="gray", justify="left", highlightthickness=0)
    entrypq_widgets[f"part0{i}"].grid(row=i+9, column=1, padx=5, pady=1, sticky="w")
    entrypq_widgets[f"part0{i}"].placeholder = f"Porque {i}"
    entrypq_widgets[f"part0{i}"].insert(tk.END, entrypq_widgets[f"part0{i}"].placeholder)
    entrypq_widgets[f"part0{i}"].bind("<FocusIn>", restore_placeholder)
    entrypq_widgets[f"part0{i}"].bind("<FocusOut>", show_placeholder)

# Criando a Página 2
tab2 = tk.Frame(tab_control)
tab_control.add(tab2, text='Página 2')

main_frame2 = tk.Frame(tab2)
main_frame2.pack()

part1_frame2 = tk.Frame(main_frame2, width=300, padx=10, pady=10)
part1_frame2.grid(row=0, column=0, sticky="nsew")

entry_width = 30
# Ações de Conteção Imediata
labels = ['Item', 'Atividade', 'Responsável', 'P. Inicial', 'P. Final']
for col, label_text in enumerate(labels):
    tk.Label(part1_frame2, text=label_text).grid(row=0, column=col, padx=5, pady=5)

entradas = {}
for row, label_text in enumerate(['input', 'input', 'input', 'input', 'input'], start=1):
    for col in range(len(labels)):
        entrada = tk.Entry(part1_frame2, width=entry_width)
        entrada.grid(row=row, column=col, padx=5, pady=5)
        entradas[(row, col)] = entrada

for row in range(1, 6):  # Percorre as linhas
    for col in range(3, 5):  # Percorre as duas últimas colunas (índices 3 e 4)
        entradas[(row, col)].bind("<KeyRelease>", formatar_data)


# Perguntas adicionais
tk.Label(part1_frame2, text="Mudanças ou Alterações Recentes: ").grid(row=8, column=0, padx=5, pady=5, sticky="w")
mudancas_input = tk.StringVar()
mudancas_input.set("")
mudancas_input = ttk.Combobox(part1_frame2,width=30, values=["", "Sim", "Não"])
mudancas_input.grid(row=8, column=1, padx=5, pady=5, sticky="w")



tk.Label(part1_frame2, text="Algum modo de falha semelhante?").grid(row=9, column=0, padx=5, pady=5, sticky="w")
falha_input = tk.StringVar()
falha_input.set("")
falha_input = ttk.Combobox(part1_frame2,width=30, values=["", "Sim", "Não"])
falha_input.grid(row=9, column=1, padx=5, pady=5, sticky="w")

# Coleta de Dados
tk.Label(part1_frame2, text="Há necessidade de Coleta de Dados?").grid(row=10, column=0, padx=5, pady=5, sticky="w")
coleta_combobox = ttk.Combobox(part1_frame2,width=30, values=["Não", "Sim", "Folha de Verificação", "Histograma", "Diagrama de Pareto", "Diagrama de Causa Raiz", "Outro"])
coleta_combobox.grid(row=10, column=1, padx=5, pady=5, sticky="w")

# Observações
tk.Label(part1_frame2, text="Observações?").grid(row=11, column=0, padx=5, pady=5, sticky="w")
observacoes_input = tk.Text(part1_frame2, width=35, height=5, font=("Arial", 9))
observacoes_input.grid(row=11, column=1, padx=5, pady=5, sticky="w")
observacoes_input.bind("<Tab>", focus_next_widget)

# Brainstorming
tk.Label(part1_frame2, text="Brainstorming:").grid(row=12, column=0, padx=5, pady=5, sticky="w")
brainstorming_input = tk.StringVar()
brainstorming_input.set("")
brainstorming_input = ttk.Combobox(part1_frame2,width=30, values=["", "Sim", "Não"])
brainstorming_input.grid(row=12, column=1, padx=5, pady=5, sticky="w")

# Diagrama de Causa e Efeito
tk.Label(part1_frame2, text="Diagrama de Causa e Efeito:").grid(row=13, column=0, padx=5, pady=5, sticky="w")
causa_efeito_input = tk.StringVar()
causa_efeito_input.set("")
causa_efeito_input = ttk.Combobox(part1_frame2,width=30, values=["", "Sim", "Não"])
causa_efeito_input.grid(row=13, column=1, padx=5, pady=5, sticky="w")

# Técnica dos 5 Porquês
tk.Label(part1_frame2, text="Técnica dos 5 Porquês:").grid(row=14, column=0, padx=5, pady=5, sticky="w")
cinco_porques_input = tk.StringVar()
cinco_porques_input.set("")
cinco_porques_input = ttk.Combobox(part1_frame2,width=30, values=["", "Sim", "Não"])
cinco_porques_input.grid(row=14, column=1, padx=5, pady=5, sticky="w")

# Causas Básicas
tk.Label(part1_frame2, text="Causas Básicas:").grid(row=15, column=0, padx=5, pady=5, sticky="w")
causas_basicas_input = tk.Entry(part1_frame2,width=40)
causas_basicas_input.grid(row=15, column=1, padx=5, pady=5, sticky="w")


# Página 3 

tab3 = tk.Frame(tab_control)
tab_control.add(tab3, text='Página 3')

# Criando o frame principal dentro da aba
main_frame3 = tk.Frame(tab3)
main_frame3.pack()

# Criando o frame específico para os elementos
part1_frame3 = tk.Frame(main_frame3, width=1000, height=500, padx=15, pady=15)
part1_frame3.pack()


# Labels para "Tipo Ação"
tk.Label(part1_frame3, text="Tipo Ação: ").place(x=0, y=0)
tipoacao_combobox = ttk.Combobox(part1_frame3, values=['Corretiva', 'Preventiva', 'Melhoria'])
tipoacao_combobox.place(x=70, y=0)

tk.Label(part1_frame3, text="As pessoas envolvidas foram comunicadas: ").place(x=0, y=30)
comunicadas_input = tk.StringVar()
comunicadas_input.set("")
comunicadas_combobox = ttk.Combobox(part1_frame3, width=30, values=["", "Sim", "Não"])
comunicadas_combobox.place(x=260, y=30)


# Definição dos rótulos
novos_labels = ['Item', 'Tipo1', 'Descrição da Ação', 'Responsável', 'P. Inicial', 'P. Final']

# Rótulos
for col, label_text in enumerate(novos_labels):
    tk.Label(part1_frame3, text="Item").place(x=20, y=85)
    tk.Label(part1_frame3, text="Tipo").place(x=110, y=85)
    tk.Label(part1_frame3, text="Descrição da Ação").place(x=270, y=85)
    tk.Label(part1_frame3, text="Responsável").place(x=500, y=85)
    tk.Label(part1_frame3, text="P. Inicial").place(x=670, y=85)
    tk.Label(part1_frame3, text="P. Final").place(x=770, y=85)

# Entradas para os novos elementos
entradas1 = {}
for row in range(5, 9):  # Percorre as linhas dos novos elementos (adiciona 2 linhas)
    for col, label_text in enumerate(['input1', 'input2', 'input3', 'input4', 'input5', 'input6']):
        x_position = 15 + col * 150
        y_position = 110 + (row - 5) * 40  # Aumentei o espaçamento vertical
        
        if label_text == 'input1':  # Para as colunas de input
            entrada1 = tk.Entry(part1_frame3, width=5)
            entrada1.place(x=x_position, y=y_position)
            entradas1[(row, col)] = entrada1
        elif label_text == 'input2':  # Para as colunas de input
            entrada1 = ttk.Combobox(part1_frame3, values=['AC', 'AP', 'ME'], width=17)
            entrada1.place(x=65, y=y_position)
        elif label_text == 'input3':  # Para as colunas de input
            entrada1 = tk.Text(part1_frame3, width=30, height=3, font=("Arial", 9))
            entrada1.bind("<Tab>", focus_next_widget)
            entrada1.place(x=220, y=y_position)
        elif label_text == 'input4':
            entrada1 = tk.Entry(part1_frame3, width=28)
            entrada1.place(x=450, y=y_position)
        elif label_text == 'input5':  # Para as colunas de input
            entrada1 = tk.Entry(part1_frame3, width=10)
            entrada1.place(x=660, y=y_position)
            entrada1.bind("<KeyRelease>", formatar_data)
        elif label_text == 'input6':  # Para as colunas de input
            entrada1 = tk.Entry(part1_frame3, width=10)
            entrada1.place(x=750, y=y_position)
            entrada1.bind("<KeyRelease>", formatar_data)
        entradas1[(row, col)] = entrada1
# Inicialize listas vazias para armazenar as informações de cada coluna


################################################################################################################################

        
         
# Labels para "Foram Executados Testes?"
tk.Label(part1_frame3, text="Foram Executados Testes?").place(x=0, y=300)
testes_input = tk.StringVar()
testes_input.set("")
testes_input = ttk.Combobox(part1_frame3,width=30, values=["", "Sim", "Não"])
testes_input.place(x=165, y=300)

# Labels e combobox para "Precisa de Nova Coleta de Dados?"
tk.Label(part1_frame3, text="Precisa de Nova Coleta de Dados?").place(x=0, y=350)
coleta_combobox_2 = ttk.Combobox(part1_frame3, width=30, values=["Não", "Sim", "Folha de Verificação", "Histograma", "Diagrama de Pareto", "Diagrama de Causa Raiz", "Outro"])
coleta_combobox_2.place(x=210, y=350)


# Página 4

tab4 = tk.Frame(tab_control)
tab_control.add(tab4, text='Página 4')

main_frame4 = tk.Frame(tab4)
main_frame4.pack()

part1_frame4 = tk.Frame(main_frame4, width=300, padx=15, pady=15)
part1_frame4.grid(row=0, column=0, sticky="nsew")



# O Problema foi Resolvido?
tk.Label(part1_frame4, text="O Problema foi Resolvido?").grid(row=0, column=0, padx=15, pady=15, sticky="w")
problema_resolvido_input = tk.Text(part1_frame4, width=50, height=7, font=("Arial", 9))
problema_resolvido_input.bind("<Tab>", focus_next_widget)
problema_resolvido_input.grid(row=0, column=1, padx=15, pady=15, sticky="w")

# Conclusão
tk.Label(part1_frame4, text="Conclusão:").grid(row=1, column=0, padx=15, pady=15, sticky="w")
conclusao_var4 = tk.StringVar()
conclusao_var4.set("")
conclusao_menu4 = ttk.Combobox(part1_frame4, values=["", "Eficaz", "Ineficaz"], textvariable=conclusao_var4)
conclusao_menu4.grid(row=1, column=1, padx=15, pady=15, sticky="w")

# Participantes
tk.Label(part1_frame4, text="Participantes:").grid(row=2, column=0, padx=15, pady=15, sticky="w")
entry_widgets4 = {}
for i in range(1, 5):
    entry_widgets4[f"part{i}"] = tk.Entry(part1_frame4)
    entry_widgets4[f"part{i}"].placeholder = f"Participante {i}"
    entry_widgets4[f"part{i}"].insert(tk.END, entry_widgets4[f"part{i}"].placeholder)
    entry_widgets4[f"part{i}"].bind("<FocusIn>", restore_placeholder)
    entry_widgets4[f"part{i}"].bind("<FocusOut>", show_placeholder)
    entry_widgets4[f"part{i}"].grid(row=1+i, column=1, padx=15, pady=5, sticky="w")

# Analise de Encerramento
tk.Label(part1_frame4, text="Análise de Encerramento:").grid(row=6, column=0, padx=15, pady=15, sticky="w")
analise_encerramento_input = tk.Text(part1_frame4, width=50,height=7, font=("Arial", 9))
analise_encerramento_input.bind("<Tab>", focus_next_widget)
analise_encerramento_input.grid(row=6, column=1, padx=15, pady=15, sticky="w")





salvar_button = tk.Button(root, text="Salvar no Excel", command=salvar_excel)
salvar_button.pack()
apagar_button = tk.Button(root, text="Apagar Todas as Informações", command=apagar_tudo)
apagar_button.pack()

root.mainloop()
